import tkinter as tk
import tkinter.messagebox
import random
from openpyxl import load_workbook


def onOK():
    if format(Entry.get()) == answer:
        tkinter.messagebox.showinfo(title='Hi', message='答對')
    else:
        tkinter.messagebox.showinfo(title='Hi', message='答錯')

window = tk.Tk()
window.title('vocabulary memorizing')

window.geometry('1600x1200')
voc = tk.Label(window, text='背單字囉Q_Q', bg='orange', font=('Arial', 12), width=30, height=2)
voc.pack()

wb = load_workbook('voc test.xlsx')
print (wb.sheetnames)
sheet = wb['工作表1']
QuizList = []
answer_sheet = []
# answer_index = 0

rows_number = sheet.max_row
number_of_sheet = random.randint(1,rows_number)
QuizList.append(number_of_sheet)
        

#獲取最大欄位數
a=0
    
quiz_number = "A"+ str(QuizList[0])
quest = sheet[quiz_number]
quiz_for_player = quest.value
quiz_for_player_onwindow = tk.Label(window, text="題目:" + quiz_for_player)
quiz_for_player_onwindow.pack()
# 將單字問項置於視窗
print (quest.value)
    # 獲取單字
pos_number = "B" + str(QuizList[0])
pos = sheet[pos_number]
pos_for_player = pos.value
pos_for_player_onwindow = tk.Label(window, text="詞性:" + pos_for_player)
pos_for_player_onwindow.pack()
print (pos.value)
# 獲取詞性
chinese_answer = "C" + str(QuizList[0])
ca = sheet[chinese_answer]
answer = ca.value
answer_sheet.append(answer)
print (ca.value)
# 獲取答案，置於答案表單# 輸入欄位
Entry = tk.Entry(window,     # 輸入欄位所在視窗
             width = 20) # 輸入欄位的寬度
Entry.pack()
# 按鈕
button = tk.Button(window, text = "OK", command = onOK)
button.pack()

window.mainloop()
